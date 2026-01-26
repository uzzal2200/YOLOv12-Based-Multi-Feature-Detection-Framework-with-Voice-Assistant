# import cv2
# import os
# import openpyxl
# import time
# import numpy as np
# import sys

# # Parse command line arguments
# video_file = None
# if len(sys.argv) > 2 and sys.argv[1] == '--video':
#     video_file = sys.argv[2]

# # Resolve all paths relative to this script's directory
# BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# # DEFAULT VIDEO PATH (hardcoded)
# DEFAULT_VIDEO = r"D:\Research\Research\Real time object detection feedback with vedio\Object-text-detection-for-visually-impaired\video_2026-01-21_22-08-00.mp4"

# # If video_file not provided via command line, use default
# if not video_file:
#     video_file = DEFAULT_VIDEO

# # If video_file provided, make it absolute path
# if video_file:
#     if not os.path.isabs(video_file):
#         video_file = os.path.join(BASE_DIR, video_file)
    
#     if os.path.exists(video_file):
#         cam = cv2.VideoCapture(video_file)
#         print(f"✓ Video loaded: {video_file}")
#     else:
#         print(f"✗ Video not found: {video_file}")
#         print(f"  Falling back to webcam...")
#         cam = cv2.VideoCapture(0)
#         print("Using webcam...")
# else:
#     cam = cv2.VideoCapture(0)
#     print("Using webcam...")

# # Known faces directory setup (absolute path)
# BASE_DIR = os.path.dirname(os.path.abspath(__file__))
# known_faces_folder = os.path.join(BASE_DIR, "known_faces_folder")
# known_face_images = []
# known_names = []

# # Load OpenCV face detector
# face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')

# # Excel, log setup (absolute paths)
# log_folder = os.path.join(BASE_DIR, "movement_logs")
# if not os.path.exists(log_folder):
#     os.makedirs(log_folder)

# excel_file_path = os.path.join(BASE_DIR, "movement_log.xlsx")
# if not os.path.exists(excel_file_path):
#     wb = openpyxl.Workbook()
#     sheet = wb.active
#     sheet.title = "Movement Logs"
#     sheet.append(["Name", "Detection Time", "Date", "Confidence"])
#     wb.save(excel_file_path)

# # --- Utils ---
# def normalize_face(gray_face):
#     """Standardize face crop for matching (resize + equalize)."""
#     face_eq = cv2.equalizeHist(gray_face)
#     return cv2.resize(face_eq, (150, 150))

# def face_match_score(face1, face2):
#     """Calculate face matching score using feature-based matching with ORB"""
#     try:
#         # Ensure both faces are the same size and normalized
#         face1_eq = normalize_face(face1)
#         face2_eq = normalize_face(face2)

#         # Use ORB for keypoint detection (faster than SIFT)
#         orb = cv2.ORB_create(nfeatures=1500)
#         kp1, des1 = orb.detectAndCompute(face1_eq, None)
#         kp2, des2 = orb.detectAndCompute(face2_eq, None)

#         if des1 is None or des2 is None:
#             # Fallback to pixel comparison
#             diff = cv2.absdiff(face1_eq, face2_eq)
#             return 1 - (np.mean(diff) / 255.0)

#         # Use BFMatcher for matching
#         bf = cv2.BFMatcher(cv2.NORM_HAMMING, crossCheck=True)
#         matches = bf.match(des1, des2)
#         matches = sorted(matches, key=lambda x: x.distance)

#         # Calculate match score based on number and quality of matches
#         if len(matches) > 0:
#             good_matches = [m for m in matches[:20] if m.distance < 70]
#             match_score = len(good_matches) / 20.0
#             avg_distance = np.mean([m.distance for m in matches[:10]]) if matches else 255
#             distance_score = 1 - (avg_distance / 255.0)
#             combined = (match_score * 0.6 + distance_score * 0.4)
#             return min(combined, 1.0)
#         else:
#             return 0.0
#     except Exception:
#         return 0.0

# # Load known faces (store actual images for simple comparison)
# if os.path.exists(known_faces_folder):
#     print(f"Checking folder: {known_faces_folder}")
#     files = os.listdir(known_faces_folder)
#     print(f"Files found: {files}")

#     for filename in files:
#         if filename.lower().endswith((".jpg", ".jpeg", ".png")):
#             filepath = os.path.join(known_faces_folder, filename)
#             print(f"Loading: {filepath}")
#             image = cv2.imread(filepath)
#             if image is not None:
#                 print(f"Image loaded successfully: {image.shape}")
#                 # Convert to grayscale for face detection
#                 gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
#                 faces = face_cascade.detectMultiScale(gray, 1.1, 4)
#                 print(f"Faces detected in {filename}: {len(faces)}")

#                 if len(faces) > 0:
#                     # Take the first detected face
#                     x, y, w, h = faces[0]
#                     face_roi = gray[y:y+h, x:x+w]
#                 else:
#                     # Fallback: take centered crop if detector misses
#                     h_img, w_img = gray.shape[:2]
#                     crop_size = int(min(h_img, w_img) * 0.7)
#                     start_x = (w_img - crop_size) // 2
#                     start_y = (h_img - crop_size) // 2
#                     face_roi = gray[start_y:start_y+crop_size, start_x:start_x+crop_size]
#                     print(f"No face detected in {filename}, using centered crop")

#                 # Normalize for comparison
#                 face_roi = normalize_face(face_roi)
#                 known_face_images.append(face_roi)
#                 known_names.append(os.path.splitext(filename)[0])
#                 print(f"Added face: {os.path.splitext(filename)[0]}")
#             else:
#                 print(f"Failed to load image: {filepath}")
# else:
#     print(f"Folder not found: {known_faces_folder}")

# print(f"Loaded {len(known_face_images)} known faces: {known_names}")
# print("Instructions:")
# print("- Press 'c' to capture new face")
# print("- Press 'q' to quit")
# print("- Detection will work automatically")

# frame_count = 0
# detection_threshold = 0.12  # Lower threshold to accept closer matches
# capture_mode = False

# def capture_new_face(frame, face_roi, gray_roi):
#     """Capture and save a new face"""
#     name = input("\nEnter name for this face: ").strip()
#     if name:
#         # Save the face image
#         filename = f"{name}.jpg"
#         filepath = os.path.join(known_faces_folder, filename)

#         # Save the original face region from frame
#         y1, y2, x1, x2 = face_roi
#         face_image = frame[y1:y2, x1:x2]
#         cv2.imwrite(filepath, face_image)

#         # Add to known faces list
#         face_roi_resized = normalize_face(gray_roi)
#         known_face_images.append(face_roi_resized)
#         known_names.append(name)

#         print(f"Face saved as '{name}' in {filepath}")
#         print(f"Now {len(known_face_images)} known faces loaded")
#         return True
#     return False

# while cam.isOpened():
#     ret, frame = cam.read()
#     if not ret:
#         break

#     frame_count += 1
#     if frame_count % 2 != 0:
#         continue  # Process every 2nd frame for efficiency

#     timestamp = time.strftime("%Y-%m-%d %H:%M:%S")
#     gray_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)

#     # Detect faces in current frame
#     faces = face_cascade.detectMultiScale(gray_frame, 1.1, 4)

#     for (x, y, w, h) in faces:
#         # Extract face region
#         face_roi = gray_frame[y:y+h, x:x+w]
#         face_roi_resized = normalize_face(face_roi)

#         # Try to match with known faces
#         name = "Unknown"
#         best_match = 0
#         confidence = 0

#         for i, known_face in enumerate(known_face_images):
#             # Calculate similarity using improved face matching
#             match_score = face_match_score(face_roi_resized, known_face)

#             if match_score > best_match:
#                 best_match = match_score
#                 confidence = match_score
#                 name = known_names[i]

#         # Determine if it's a known face (using lower threshold now)
#         if confidence < detection_threshold:
#             name = "Unknown"
#             confidence = 0

#         # Draw rectangle and label
#         if name != "Unknown":
#             color = (0, 255, 0)  # Green for known faces
#             status = "KNOWN"
#             label = f"{status}: {name} ({confidence:.2f})"
#         else:
#             color = (0, 0, 255)  # Red for unknown faces
#             status = "UNKNOWN"
#             label = f"{status}"

#         cv2.rectangle(frame, (x, y), (x+w, y+h), color, 2)

#         # Add background for text
#         text_size = cv2.getTextSize(label, cv2.FONT_HERSHEY_SIMPLEX, 0.7, 2)[0]
#         cv2.rectangle(frame, (x, y-35), (x + text_size[0] + 10, y), color, -1)
#         cv2.putText(frame, label, (x+5, y-10), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 255), 2)

#         # Only log if confidence is high enough
#         if confidence > 0.5 or name == "Unknown":
#             # Log detection
#             image_path = os.path.join(log_folder, f"{name}_{timestamp.replace(':', '-')}.jpg")
#             cv2.imwrite(image_path, frame)

#             # Save to Excel with error handling
#             try:
#                 wb = openpyxl.load_workbook(excel_file_path)
#                 sheet = wb.active
#                 sheet.append([name, timestamp, time.strftime("%Y-%m-%d"), f"{confidence:.2f}"])
#                 wb.save(excel_file_path)
#             except Exception as e:
#                 print(f"Excel error: {e}")
#                 # Create new file if corrupted
#                 wb = openpyxl.Workbook()
#                 sheet = wb.active
#                 sheet.title = "Movement Logs"
#                 sheet.append(["Name", "Detection Time", "Date", "Confidence"])
#                 sheet.append([name, timestamp, time.strftime("%Y-%m-%d"), f"{confidence:.2f}"])
#                 wb.save(excel_file_path)
#                 print("Created new Excel file")

#             print(f"Detected: {name} (confidence: {confidence:.2f}) at {timestamp}")

#     cv2.imshow("Face Detection - Simple Version", frame)
#     key = cv2.waitKey(1) & 0xFF
#     if key == ord('q'):
#         break
#     elif key == ord('c'):
#         # Capture the first detected face in the current frame
#         if len(faces) > 0:
#             x, y, w, h = faces[0]
#             try:
#                 captured = capture_new_face(
#                     frame,
#                     (y, y + h, x, x + w),
#                     gray_frame[y:y + h, x:x + w]
#                 )
#                 if captured:
#                     print("New face captured and added. Try again in front of camera.")
#             except Exception as e:
#                 print(f"Capture failed: {e}")
#         else:
#             print("No face found to capture. Center your face and try again.")

# cam.release()
# cv2.destroyAllWindows()
# print("Face detection stopped.")

# # python known_unknown_detection.py




import cv2
import os
import openpyxl
import time
import numpy as np
import sys
import csv

# =====================================================
# PATH & VIDEO SETUP
# =====================================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

DEFAULT_VIDEO = r"D:\Research\Research\Real time object detection feedback with vedio\Object-text-detection-for-visually-impaired\video_2026-01-21_22-08-00.mp4"

video_file = DEFAULT_VIDEO
if len(sys.argv) > 2 and sys.argv[1] == "--video":
    video_file = sys.argv[2]

cam = cv2.VideoCapture(video_file if os.path.exists(video_file) else 0)
print("✓ Video source opened")

# =====================================================
# FOLDERS & FILES
# =====================================================
known_faces_folder = os.path.join(BASE_DIR, "known_faces_folder")
log_folder = os.path.join(BASE_DIR, "movement_logs")
os.makedirs(log_folder, exist_ok=True)

excel_file = os.path.join(BASE_DIR, "movement_log.xlsx")
eval_csv = os.path.join(BASE_DIR, "face_eval_log.csv")

# Create evaluation CSV if not exists
if not os.path.exists(eval_csv):
    with open(eval_csv, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["timestamp", "ground_truth", "predicted", "confidence"])

# Create Excel if not exists
if not os.path.exists(excel_file):
    wb = openpyxl.Workbook()
    sh = wb.active
    sh.title = "Movement Logs"
    sh.append(["Name", "Detection Time", "Date", "Confidence"])
    wb.save(excel_file)

# =====================================================
# FACE DETECTOR
# =====================================================
face_cascade = cv2.CascadeClassifier(
    cv2.data.haarcascades + "haarcascade_frontalface_default.xml"
)

# =====================================================
# UTILS
# =====================================================
def normalize_face(gray):
    gray = cv2.equalizeHist(gray)
    return cv2.resize(gray, (150, 150))

def face_match_score(f1, f2):
    orb = cv2.ORB_create(1500)
    kp1, des1 = orb.detectAndCompute(f1, None)
    kp2, des2 = orb.detectAndCompute(f2, None)
    if des1 is None or des2 is None:
        diff = cv2.absdiff(f1, f2)
        return 1 - (np.mean(diff) / 255.0)

    bf = cv2.BFMatcher(cv2.NORM_HAMMING, crossCheck=True)
    matches = bf.match(des1, des2)
    matches = sorted(matches, key=lambda x: x.distance)

    good = [m for m in matches[:20] if m.distance < 70]
    return min(len(good) / 20.0, 1.0)

# =====================================================
# LOAD KNOWN FACES
# =====================================================
known_faces = []
known_names = []

if os.path.exists(known_faces_folder):
    for f in os.listdir(known_faces_folder):
        if f.lower().endswith((".jpg", ".png", ".jpeg")):
            img = cv2.imread(os.path.join(known_faces_folder, f))
            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
            faces = face_cascade.detectMultiScale(gray, 1.1, 4)

            if len(faces) > 0:
                x, y, w, h = faces[0]
                face = normalize_face(gray[y:y+h, x:x+w])
                known_faces.append(face)
                known_names.append(os.path.splitext(f)[0])

print(f"✓ Loaded {len(known_faces)} known faces")

# =====================================================
# PARAMETERS
# =====================================================
threshold = 0.12
frame_id = 0

print("\nPress:")
print(" c → capture new face")
print(" q → quit\n")

# =====================================================
# MAIN LOOP
# =====================================================
while cam.isOpened():
    ret, frame = cam.read()
    if not ret:
        break

    frame_id += 1
    if frame_id % 2 != 0:
        continue

    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    faces = face_cascade.detectMultiScale(gray, 1.1, 4)

    timestamp = time.strftime("%Y-%m-%d %H:%M:%S")

    for (x, y, w, h) in faces:
        roi = normalize_face(gray[y:y+h, x:x+w])

        name = "Unknown"
        best = 0.0

        for i, kf in enumerate(known_faces):
            score = face_match_score(roi, kf)
            if score > best:
                best = score
                name = known_names[i]

        if best < threshold:
            name = "Unknown"
            best = 0.0

        # ==========================
        # DRAW
        # ==========================
        color = (0, 255, 0) if name != "Unknown" else (0, 0, 255)
        label = f"{name} ({best:.2f})" if name != "Unknown" else "UNKNOWN"

        cv2.rectangle(frame, (x, y), (x+w, y+h), color, 2)
        cv2.putText(frame, label, (x, y-10),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.7, color, 2)

        # ==========================
        # LOGGING (CRITICAL)
        # ==========================
        ground_truth = "Known" if name != "Unknown" else "Unknown"

        with open(eval_csv, "a", newline="") as f:
            writer = csv.writer(f)
            writer.writerow([timestamp, ground_truth, name, f"{best:.2f}"])

        if best > 0.5 or name == "Unknown":
            img_path = os.path.join(
                log_folder, f"{name}_{timestamp.replace(':','-')}.jpg"
            )
            cv2.imwrite(img_path, frame)

            wb = openpyxl.load_workbook(excel_file)
            sh = wb.active
            sh.append([name, timestamp, time.strftime("%Y-%m-%d"), f"{best:.2f}"])
            wb.save(excel_file)

        print(f"Detected: {name} | Confidence: {best:.2f}")

    cv2.imshow("Known–Unknown Face Recognition", frame)

    key = cv2.waitKey(1) & 0xFF
    if key == ord("q"):
        break

cam.release()
cv2.destroyAllWindows()
print("✓ Detection stopped")
